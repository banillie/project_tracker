import datetime

from django.contrib.auth.models import User

from openpyxl import Workbook

from engagements.models import (
    Engagement,
    EngagementTopic,
)
from projects.models import Project
from stakeholders.models import Stakeholder
from ppdds.models import PPDD


def excel_download(output: str) -> None:
    wb = Workbook()

    ws = wb.create_sheet("Engagements")  # Engagement output
    fields_all = [f.name for f in Engagement._meta.get_fields()]
    # remove = ['id',
    #           'user'
    #           ] # fields currently not required for user
    # fields = [x for x in fields_all if x not in remove]

    for x, f in enumerate(fields_all):
        ws.cell(row=1, column=x + 1).value = f.upper()
        for i, p in enumerate(Engagement.objects.all().order_by('-date')):  # order by date
            v = getattr(p, f)
            try:
                if f == 'user':
                    ws.cell(row=i + 2, column=x + 1).value = str(v)
                else:
                    ws.cell(row=i + 2, column=x + 1).value = v
                    if isinstance(v, datetime.date):
                        ws.cell(row=i + 2, column=x + 1).number_format = "dd/mm/yy"
            except ValueError:
                m2m_list = []
                attribute_list = v.all()  # for many-to-many model instance fields.
                for y in attribute_list:
                    m2m_list.append(str(y))
                ws.cell(row=i + 2, column=x + 1).value = ', '.join(m2m_list)

    ws = wb.create_sheet("Projects")  # Project output
    fields_all = [f.name for f in Project._meta.get_fields()]
    remove = [
        "engagement",
        # "id",
        # "user",
        "type",
        "slug",
        "governance",
        "stage",
        "live",
        "timestamp",
        "updated",
    ]  # fields currently not reqired for user
    fields = [x for x in fields_all if x not in remove]

    for x, f in enumerate(fields):
        ws.cell(row=1, column=x + 1).value = f.upper()
        for i, p in enumerate(Project.objects.all()):
            val = getattr(p, f)
            try:
                ws.cell(row=i + 2, column=x + 1).value = getattr(p, f)
            except ValueError:
                if val is not None:
                    ws.cell(row=i + 2, column=x + 1).value = str(getattr(p, f))
                else:
                    ''

    ws = wb.create_sheet("Stakeholders")  # Stakeholder output
    fields_all = [f.name for f in Stakeholder._meta.get_fields()]
    remove = ['engagement',
              # 'id',
              'user',
              'slug',
              'group',
              'live']  # fields currently not reqired for user
    fields = [x for x in fields_all if x not in remove]

    for x, f in enumerate(fields):
        ws.cell(row=1, column=x + 1).value = f.upper()
        for i, p in enumerate(Stakeholder.objects.all().order_by('last_name')):
            try:
                ws.cell(row=i + 2, column=x + 1).value = getattr(p, f)
            except ValueError:
                ws.cell(row=i + 2, column=x + 1).value = str(getattr(p, f))

    ws = wb.create_sheet("PPDD")  # PPDD output
    fields_all = [f.name for f in PPDD._meta.get_fields()]
    remove = ['engagement',
              # 'id',
              'user',
              'slug',
              'tele_no',
              'live']  # fields currently not reqired for user
    fields = [x for x in fields_all if x not in remove]

    for x, f in enumerate(fields):
        ws.cell(row=1, column=x + 1).value = f.upper()
        for i, p in enumerate(PPDD.objects.all().order_by('last_name')):
            try:
                ws.cell(row=i + 2, column=x + 1).value = getattr(p, f)
            except ValueError:
                ws.cell(row=i + 2, column=x + 1).value = str(getattr(p, f))

    wb.remove(wb["Sheet"])

    wb.save(output)


def excel_download_pbi(output: str) -> None:
    wb = Workbook()
    ws = wb.create_sheet("Engagements")  # Engagement output

    e_qs = Engagement.objects.raw(
        'SELECT * FROM engagements_engagement'
    )
    for x, f in enumerate(e_qs):
        ws.cell(row=x + 2, column=1).value = f.id
        ws.cell(row=x + 2, column=2).value = f.date
        ws.cell(row=x + 2, column=2).number_format = "dd/mm/yy"
        try:
            ws.cell(row=x + 2, column=3).value = f.user.id
        except AttributeError:  # not all have user attached.
            pass

    ws.cell(row=1, column=1).value = 'ENGAGEMENT_ID'
    ws.cell(row=1, column=2).value = 'DATE'
    ws.cell(row=1, column=3).value = 'USER'

    ws = wb.create_sheet('Engagement_Topics')
    et_qs = Engagement.objects.raw(
        'SELECT * FROM engagements_engagement_topics'
    )
    for x, f in enumerate(et_qs):
        ws.cell(row=x + 2, column=1).value = f.id
        ws.cell(row=x + 2, column=2).value = f.engagement_id
        ws.cell(row=x + 2, column=3).value = f.engagementtopic_id
        ws.cell(row=x + 2, column=4).value = str(EngagementTopic.objects.get(id=f.engagementtopic_id))

    ws.cell(row=1, column=1).value = 'ID'
    ws.cell(row=1, column=2).value = 'ENGAGEMENT ID'
    ws.cell(row=1, column=3).value = 'ENGAGEMENT_TOPIC ID'
    ws.cell(row=1, column=4).value = 'ENGAGEMENT_TOPIC NAME'

    ws = wb.create_sheet('Engagement_Projects')
    ep_qs = Engagement.objects.raw(
        'SELECT * FROM engagements_engagement_projects'
    )
    for x, ep in enumerate(ep_qs):
        ws.cell(row=x + 2, column=1).value = ep.id
        ws.cell(row=x + 2, column=2).value = ep.engagement_id
        ws.cell(row=x + 2, column=3).value = ep.project_id
        instance = Project.objects.get(id=ep.project_id)
        ws.cell(row=x + 2, column=4).value = str(instance)
        ws.cell(row=x + 2, column=5).value = str(instance.tier)
        ws.cell(row=x + 2, column=6).value = str(instance.sort)
        ws.cell(row=x + 2, column=7).value = str(instance.dft_group)

    ws.cell(row=1, column=1).value = 'ID'
    ws.cell(row=1, column=2).value = 'ENGAGEMENT ID'
    ws.cell(row=1, column=3).value = 'PROJECT ID'
    ws.cell(row=1, column=4).value = 'PROJECT NAME'
    ws.cell(row=1, column=5).value = 'PROJECT TIER'
    ws.cell(row=1, column=6).value = 'PROJECT TYPE'
    ws.cell(row=1, column=7).value = 'PROJECT GROUP'

    ws = wb.create_sheet("Projects")  # Project output
    projects_qs = Project.objects.raw('SELECT * FROM projects_project')
    for x, f in enumerate(projects_qs):
        ws.cell(row=x + 2, column=1).value = f.abbreviation
        ws.cell(row=x + 2, column=2).value = str(f.dft_group)
        ws.cell(row=x + 2, column=3).value = f.id
        ws.cell(row=x + 2, column=4).value = f.name
        ws.cell(row=x + 2, column=5).value = str(f.stage_name)
        ws.cell(row=x + 2, column=6).value = str(f.tier)
        ws.cell(row=x + 2, column=7).value = str(f.sort)

    ws.cell(row=1, column=1).value = 'ABBREVIATION'
    ws.cell(row=1, column=2).value = 'DFT_GROUP'
    ws.cell(row=1, column=3).value = 'ID'
    ws.cell(row=1, column=4).value = 'NAME'
    ws.cell(row=1, column=5).value = 'STAGE_NAME'
    ws.cell(row=1, column=6).value = 'TIER'
    ws.cell(row=1, column=7).value = 'TYPE'

    ## old method
    # fields_all = [f.name for f in Project._meta.get_fields()]
    # remove = [
    #     "engagement",
    #     # "id",
    #     "user",
    #     "slug",
    #     "governance",
    #     "stage",
    #     "live",
    #     "timestamp",
    #     "updated",
    #     "scope"
    # ]  # fields currently not required for user
    # fields = [x for x in fields_all if x not in remove]
    #
    # for x, f in enumerate(fields):
    #     ws.cell(row=1, column=x + 1).value = f.upper()
    #     for i, p in enumerate(Project.objects.all()):
    #         val = getattr(p, f)
    #         try:
    #             ws.cell(row=i + 2, column=x + 1).value = getattr(p, f)
    #         except ValueError:
    #             if val is not None:
    #                 ws.cell(row=i + 2, column=x + 1).value = str(getattr(p, f))
    #             else:
    #                 pass

    ws = wb.create_sheet("PPDD")  # PPDD output
    ppdd_qs = PPDD.objects.raw('SELECT * FROM ppdds_ppdd')
    for x, f in enumerate(ppdd_qs):
        ws.cell(row=x + 2, column=1).value = f.id
        ws.cell(row=x + 2, column=2).value = f.first_name + ' ' + f.last_name
        ws.cell(row=x + 2, column=3).value = str(f.division)
        ws.cell(row=x + 2, column=4).value = f.team
        ws.cell(row=x + 2, column=5).value = f.role

    ws.cell(row=1, column=1).value = 'ID'
    ws.cell(row=1, column=2).value = 'NAME'
    ws.cell(row=1, column=3).value = 'DIVISION'
    ws.cell(row=1, column=4).value = 'TEAM'
    ws.cell(row=1, column=5).value = 'ROLE'

    ## previous method for extracting data.
    # fields_all = [f.name for f in PPDD._meta.get_fields()]
    # remove = ['engagement',
    #           'id',
    #           'user',
    #           'slug',
    #           'tele_no',
    #           'live']  # fields currently not reqired for user
    # fields = [x for x in fields_all if x not in remove]
    #
    # for x, f in enumerate(fields):
    #     ws.cell(row=1, column=x + 1).value = f.upper()
    #     for i, p in enumerate(PPDD.objects.all().order_by('last_name')):
    #         try:
    #             ws.cell(row=i + 2, column=x + 1).value = getattr(p, f)
    #         except ValueError:
    #             ws.cell(row=i + 2, column=x + 1).value = str(getattr(p, f))

    ws = wb.create_sheet("User")
    u_qs = User.objects.raw(
        'SELECT * FROM auth_user'
    )
    for x, f in enumerate(u_qs):
        ws.cell(row=x + 2, column=1).value = f.id
        user_name = f.first_name + ' ' + f.last_name
        ws.cell(row=x + 2, column=2).value = user_name
        ## work around until extend user model with custom fields.
        for v in ppdd_qs:
            ppdd_name = v.first_name + ' ' + v.last_name
            if ppdd_name == user_name:
                ws.cell(row=x + 2, column=3).value = str(v.division)

    ws.cell(row=1, column=1).value = 'USER_ID'
    ws.cell(row=1, column=2).value = 'USERNAME'
    ws.cell(row=1, column=3).value = 'DIVISION'

    ws = wb.create_sheet("Stakeholders")  # Stakeholder output
    fields_all = [f.name for f in Stakeholder._meta.get_fields()]
    remove = ['engagement',
              'id',
              'user',
              'slug',
              'group',
              'live']  # fields currently not reqired for user
    fields = [x for x in fields_all if x not in remove]

    for x, f in enumerate(fields):
        ws.cell(row=1, column=x + 1).value = f.upper()
        for i, p in enumerate(Stakeholder.objects.all().order_by('last_name')):
            try:
                ws.cell(row=i + 2, column=x + 1).value = getattr(p, f)
            except ValueError:
                ws.cell(row=i + 2, column=x + 1).value = str(getattr(p, f))

    wb.remove(wb["Sheet"])

    wb.save(output)
