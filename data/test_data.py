import os

from project_tracker.settings import ROOT_DIR
from engagements.models import Engagement, EngagementType, EngagementWorkStream
from projects.models import Project
from stakeholders.models import Stakeholder, StakeholderOrg
from ppdds.models import PPDD

from openpyxl import load_workbook


def upload_data():
    data_path = os.path.join(ROOT_DIR, 'project_tracker/data/project_tracker_data.xlsx')
    wb = load_workbook(data_path)

    # Stakeholders
    ws = wb["Stakeholders"]
    org_list = []
    for row in range(2, ws.max_row+1):
        org = ws.cell(row=row, column=3).value
        if org not in org_list:
            org_list.append(org)
        else:
            pass

    for x in org_list:
        StakeholderOrg.objects.create(org=x)

    all_entries = {}
    for row in range(2, ws.max_row+1):
        single_entry = {}
        single_entry["first_name"] = ws.cell(row=row, column=1).value.strip()
        # using strip to tidy string whitespacing
        single_entry["last_name"] = ws.cell(row=row, column=2).value.strip()
        org = ws.cell(row=row, column=3).value
        single_entry["organisation"] = StakeholderOrg.objects.get(org=org)
        single_entry["group"] = ws.cell(row=row, column=4).value
        single_entry["team"] = ws.cell(row=row, column=5).value
        single_entry["role"] = ws.cell(row=row, column=6).value
        single_entry["tele_no"] = ws.cell(row=row, column=7).value
        all_entries[int(row)] = single_entry

    for x in all_entries:
        Stakeholder.objects.create(
            **all_entries[x],
        )

    # PPDD data
    ws = wb['PPDDs']
    all_entries = {}
    for row in range(2, ws.max_row+1):
        single_entry = {}
        single_entry["first_name"] = ws.cell(row=row, column=1).value.strip()
        single_entry["last_name"] = ws.cell(row=row, column=2).value.strip()
        single_entry["role"] = ws.cell(row=row, column=3).value
        single_entry["team"] = ws.cell(row=row, column=4).value
        single_entry["tele_no"] = ws.cell(row=row, column=5).value
        all_entries[int(row)] = single_entry

    for x in all_entries:
        PPDD.objects.create(
            **all_entries[x],
        )

    # Project data
    ws = wb['Projects']

    all_entries = {}
    for row in range(2, ws.max_row+1):
        single_entry = {}
        single_entry["name"] = ws.cell(row=row, column=2).value
        single_entry["type"] = ws.cell(row=row, column=1).value
        single_entry["abbreviation"] = ws.cell(row=row, column=3).value
        single_entry["governance"] = ws.cell(row=row, column=4).value
        all_entries[int(row)] = single_entry

    for x in all_entries:
        Project.objects.create(
            **all_entries[x],
        )

    # Engagement data
    ws = wb['EngagementTypes']
    type_list = []
    for row in range(2, ws.max_row+1):
        type = ws.cell(row=row, column=1).value
        if type:
            if type not in type_list:
                type_list.append(type)
            else:
                pass
        else:
            pass

    for x in type_list:
        EngagementType.objects.create(type=x)

    wsheet = wb['EngagementWorkStreams']
    ws_list = []
    for row in range(2, wsheet.max_row+1):
        ws = wsheet.cell(row=row, column=1).value
        if ws not in ws_list:
            ws_list.append(ws)
        else:
            pass

    for x in ws_list:
        EngagementWorkStream.objects.create(work_stream=x)

    ws = wb['Engagements']
    for row in range(2, ws.max_row+1):
        eng = Engagement(
            date=ws.cell(row=row, column=1).value,
            summary=ws.cell(row=row, column=7).value,

        )
        eng.save()

        project_list = ws.cell(row=row, column=2).value.split(", ")
        p_object_list = []
        for name in project_list:
            p_object_list.append(Project.objects.get(name=name))
        eng.projects.add(*p_object_list)

        stakeholder_list = ws.cell(row=row, column=4).value.split(", ")
        object_list = []
        for name in stakeholder_list:
            split_name = name.split(" ")
            object_list.append(Stakeholder.objects.get(
                first_name=split_name[0],
                last_name=split_name[1])
            )
        eng.stakeholders.add(*object_list)

        ppdd_list = ws.cell(row=row, column=3).value.split(", ")
        object_list = []
        for name in ppdd_list:
            split_name = name.split(" ")
            object_list.append(PPDD.objects.get(
                first_name=split_name[0],
                last_name=split_name[1])
            )
        eng.ppdds.add(*object_list)

        type_list = ws.cell(row=row, column=5).value.split(", ")
        type_object_list = []
        for type in type_list:
            type_object_list.append(EngagementType.objects.get(type=type))
        eng.engagement_types.add(*type_object_list)

        wstream = ws.cell(row=row, column=6).value
        if wstream:
            ws_list = wstream.value.split(", ")
            ws_object_list = []
            for x in ws_list:
                ws_object_list.append(EngagementWorkStream.objects.get(work_stream=x))
            eng.engagement_workstreams.add(*ws_object_list)
        else:
            pass
